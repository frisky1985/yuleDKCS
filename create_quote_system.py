import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()

# ============ 定义样式 ============
# 标题样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

# 子标题样式
sub_header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
sub_header_font = Font(bold=True, size=10)

# 汇总表高亮
summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
summary_font = Font(bold=True, color="006100")

# 金额格式样式
money_font = Font(color="0000FF")  # 蓝色表示可输入
formula_font = Font(color="000000")  # 黑色表示公式
link_font = Font(color="008000")  # 绿色表示跨表引用

# 边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============ 1. 报价范围表 ============
ws_scope = wb.active
ws_scope.title = "报价范围"

# 标题
ws_scope['A1'] = "报价范围表"
ws_scope['A1'].font = Font(bold=True, size=14)
ws_scope.merge_cells('A1:D1')

# 表头
headers_scope = ["项目类型", "报价范围（元）", "备注", "适用场景"]
for col, header in enumerate(headers_scope, 1):
    cell = ws_scope.cell(row=3, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# 数据
scope_data = [
    ["小型项目", "5,000 - 20,000", "基础功能，1-2周交付", "简单网站、小程序"],
    ["中型项目", "20,000 - 80,000", "标准功能，2-6周交付", "企业官网、管理系统"],
    ["大型项目", "80,000 - 300,000", "复杂功能，1-3月交付", "电商平台、定制系统"],
    ["企业级项目", "300,000+", "高度定制，3月+交付", "大型平台、集成系统"],
]

for row_idx, row_data in enumerate(scope_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_scope.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

# 列宽
ws_scope.column_dimensions['A'].width = 15
ws_scope.column_dimensions['B'].width = 20
ws_scope.column_dimensions['C'].width = 25
ws_scope.column_dimensions['D'].width = 25

# ============ 2. 明细表 ============
ws_detail = wb.create_sheet("报价明细")

# 标题
ws_detail['A1'] = "项目报价明细表"
ws_detail['A1'].font = Font(bold=True, size=14)
ws_detail.merge_cells('A1:K1')

# 说明
ws_detail['A2'] = "说明：蓝色字体为可编辑项，黑色为自动计算"
ws_detail['A2'].font = Font(italic=True, color="666666", size=9)
ws_detail.merge_cells('A2:K2')

# 表头 - 分为几个部分
detail_headers = [
    "序号", "项目名称", "项目类型", 
    "工时投入", "", "",  # 工时部分
    "硬件投入", "",      # 硬件部分
    "制造投入", "",      # 制造部分
    "小计（元）"
]

sub_headers = [
    "", "", "",
    "人工（人天）", "单价（元/天）", "金额（元）",
    "设备名称", "金额（元）",
    "材料/工艺", "金额（元）",
    ""
]

# 主表头 - 先设置值，再合并
for col, header in enumerate(detail_headers, 1):
    cell = ws_detail.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# 合并单元格（工时投入 D-F, 硬件投入 G-H, 制造投入 I-J）
ws_detail.merge_cells('D4:F4')
ws_detail.merge_cells('G4:H4')
ws_detail.merge_cells('I4:J4')

# 子表头
for col, header in enumerate(sub_headers, 1):
    cell = ws_detail.cell(row=5, column=col, value=header)
    cell.fill = sub_header_fill
    cell.font = sub_header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# 示例数据行（3行示例）
example_data = [
    [1, "网站前端开发", "中型项目", 15, 800, "=D6*E6", "云服务器", 3000, "UI设计外包", 5000, "=F6+H6+J6"],
    [2, "后端API开发", "中型项目", 20, 1000, "=D7*E7", "数据库服务", 2000, "第三方接口", 3000, "=F7+H7+J7"],
    [3, "系统测试部署", "小型项目", 5, 600, "=D8*E8", "测试设备", 1000, "部署服务", 1500, "=F8+H8+J8"],
]

for row_idx, row_data in enumerate(example_data, 6):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        
        # 设置字体颜色
        if col_idx in [4, 5, 7, 9]:  # 可输入项
            cell.font = money_font
        elif col_idx in [6, 8, 10, 11]:  # 计算项
            cell.font = formula_font

# 合计行
total_row = 9
ws_detail.cell(row=total_row, column=2, value="合计").font = Font(bold=True)
ws_detail.cell(row=total_row, column=2).fill = summary_fill

# 工时合计
ws_detail.cell(row=total_row, column=4, value="=SUM(D6:D8)")
ws_detail.cell(row=total_row, column=6, value="=SUM(F6:F8)")

# 硬件合计
ws_detail.cell(row=total_row, column=8, value="=SUM(H6:H8)")

# 制造合计
ws_detail.cell(row=total_row, column=10, value="=SUM(J6:J8)")

# 总计
ws_detail.cell(row=total_row, column=11, value="=SUM(K6:K8)")

for col in range(1, 12):
    cell = ws_detail.cell(row=total_row, column=col)
    cell.border = thin_border
    cell.fill = summary_fill
    cell.font = Font(bold=True)

# 列宽设置
col_widths = [6, 20, 12, 12, 14, 12, 15, 12, 15, 12, 14]
for idx, width in enumerate(col_widths, 1):
    ws_detail.column_dimensions[get_column_letter(idx)].width = width

# 行高
for row in range(4, 10):
    ws_detail.row_dimensions[row].height = 22

# ============ 3. 汇总表 ============
ws_summary = wb.create_sheet("报价汇总")

# 标题
ws_summary['A1'] = "项目报价汇总表"
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary.merge_cells('A1:D1')

# 说明
ws_summary['A2'] = "本表自动汇总明细表数据"
ws_summary['A2'].font = Font(italic=True, color="008000", size=9)
ws_summary.merge_cells('A2:D2')

# 表头
summary_headers = ["费用类别", "金额（元）", "占比", "备注"]
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# 汇总数据（引用明细表）
summary_data = [
    ["一、工时费用", "='报价明细'!F9", "=B5/B8", "人工投入"],
    ["二、硬件费用", "='报价明细'!H9", "=B6/B8", "设备/服务采购"],
    ["三、制造费用", "='报价明细'!J9", "=B7/B8", "材料/外包服务"],
    ["总计", "='报价明细'!K9", "100%", ""],
]

for row_idx, row_data in enumerate(summary_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        
        if row_idx == 8:  # 总计行
            cell.fill = summary_fill
            cell.font = Font(bold=True, color="006100")
        else:
            if col_idx == 2:
                cell.font = link_font
            elif col_idx == 3:
                cell.number_format = '0.0%'

# 添加报价分析
ws_summary['A10'] = "报价分析"
ws_summary['A10'].font = Font(bold=True, size=12)

analysis_items = [
    ["项目数量", "=COUNTA('报价明细'!B6:B20)", "个"],
    ["平均单项金额", "=IF(B11>0,B8/B11,0)", "元"],
    ["最大单项", "=MAX('报价明细'!K6:K20)", "元"],
    ["最小单项", "=MIN('报价明细'!K6:K20)", "元"],
]

for idx, (label, formula, unit) in enumerate(analysis_items, 11):
    ws_summary.cell(row=idx, column=1, value=label).border = thin_border
    ws_summary.cell(row=idx, column=2, value=formula).border = thin_border
    ws_summary.cell(row=idx, column=2).font = formula_font
    ws_summary.cell(row=idx, column=3, value=unit).border = thin_border

# 列宽
ws_summary.column_dimensions['A'].width = 18
ws_summary.column_dimensions['B'].width = 18
ws_summary.column_dimensions['C'].width = 12
ws_summary.column_dimensions['D'].width = 20

# 保存文件
output_path = r"C:\Users\Lenovo\.qclaw\workspace\自动报价系统.xlsx"
wb.save(output_path)
print(f"[OK] 报价系统已创建: {output_path}")
